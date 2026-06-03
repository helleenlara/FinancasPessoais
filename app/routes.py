from flask import Blueprint, render_template, redirect, url_for, flash, request
from flask_login import login_required, current_user
from app import db
from app.models import Conta, Lancamento, Cofre, Transferencia, CartaoCredito, CompraCartao, GastoFixo
from datetime import datetime, date
import csv
import io

main = Blueprint('main', __name__)

# ─── DASHBOARD ───────────────────────────────────────────────
@main.route('/dashboard')
@login_required
def dashboard():
    contas = Conta.query.filter_by(usuario_id=current_user.id).all()
    cofres = Cofre.query.filter_by(usuario_id=current_user.id).all()

    hoje = date.today()
    lancamentos_mes = Lancamento.query.filter_by(usuario_id=current_user.id).filter(
        db.extract('month', Lancamento.data) == hoje.month,
        db.extract('year', Lancamento.data) == hoje.year
    ).order_by(Lancamento.data.desc()).all()

    total_entradas = sum(l.valor for l in lancamentos_mes if l.tipo == 'entrada')
    total_saidas = sum(l.valor for l in lancamentos_mes if l.tipo == 'saida')
    saldo_total = sum(c.saldo_atual for c in contas)
    ultimos = lancamentos_mes[:5]

    return render_template('dashboard.html',
        contas=contas, cofres=cofres,
        total_entradas=total_entradas,
        total_saidas=total_saidas,
        saldo_total=saldo_total,
        ultimos=ultimos
    )

# ─── CONTAS ──────────────────────────────────────────────────
@main.route('/contas')
@login_required
def contas():
    contas = Conta.query.filter_by(usuario_id=current_user.id).all()
    return render_template('contas.html', contas=contas)

@main.route('/contas/nova', methods=['GET', 'POST'])
@login_required
def nova_conta():
    if request.method == 'POST':
        conta = Conta(
            nome=request.form['nome'],
            tipo=request.form['tipo'],
            saldo_inicial=float(request.form.get('saldo_inicial', 0)),
            usuario_id=current_user.id
        )
        db.session.add(conta)
        db.session.commit()
        flash('Conta criada!', 'success')
        return redirect(url_for('main.contas'))
    return render_template('contas.html', modo='nova',
                           contas=Conta.query.filter_by(usuario_id=current_user.id).all())

@main.route('/contas/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_conta(id):
    conta = Conta.query.filter_by(id=id, usuario_id=current_user.id).first_or_404()
    if request.method == 'POST':
        conta.nome = request.form['nome']
        conta.tipo = request.form['tipo']
        conta.saldo_inicial = float(request.form.get('saldo_inicial', 0))
        db.session.commit()
        flash('Conta atualizada!', 'success')
        return redirect(url_for('main.contas'))
    return render_template('contas.html', modo='editar', conta_editar=conta,
                           contas=Conta.query.filter_by(usuario_id=current_user.id).all())

@main.route('/contas/excluir/<int:id>', methods=['POST'])
@login_required
def excluir_conta(id):
    conta = Conta.query.filter_by(id=id, usuario_id=current_user.id).first_or_404()
    db.session.delete(conta)
    db.session.commit()
    flash('Conta excluída.', 'info')
    return redirect(url_for('main.contas'))

# ─── LANÇAMENTOS ─────────────────────────────────────────────
@main.route('/lancamentos')
@login_required
def lancamentos():
    contas = Conta.query.filter_by(usuario_id=current_user.id).all()
    cartoes = CartaoCredito.query.filter_by(usuario_id=current_user.id).all()
    todos = Lancamento.query.filter_by(usuario_id=current_user.id).order_by(Lancamento.data.desc()).all()
    return render_template('lancamentos.html', lancamentos=todos, contas=contas,
                           cartoes=cartoes,
                           categorias_entrada=Lancamento.CATEGORIAS_ENTRADA,
                           categorias_saida=Lancamento.CATEGORIAS_SAIDA)

@main.route('/lancamentos/novo', methods=['POST'])
@login_required
def novo_lancamento():
    origem = request.form['conta_id']
    descricao = request.form['descricao']
    valor = float(request.form['valor'])
    tipo = request.form['tipo']
    categoria = request.form['categoria']
    data = datetime.strptime(request.form['data'], '%Y-%m-%d').date()

    if origem.startswith('cartao_'):
        cartao_id = int(origem.replace('cartao_', ''))
        CartaoCredito.query.filter_by(id=cartao_id, usuario_id=current_user.id).first_or_404()
        compra = CompraCartao(
            descricao=descricao,
            valor=valor,
            categoria=categoria,
            data=data,
            parcelas=1,
            cartao_id=cartao_id,
            usuario_id=current_user.id
        )
        db.session.add(compra)
        db.session.commit()
        flash('Lançamento registrado no cartão!', 'success')
    else:
        conta_id = int(origem.replace('conta_', ''))
        Conta.query.filter_by(id=conta_id, usuario_id=current_user.id).first_or_404()
        lancamento = Lancamento(
            descricao=descricao,
            valor=valor,
            tipo=tipo,
            categoria=categoria,
            data=data,
            conta_id=conta_id,
            usuario_id=current_user.id
        )
        db.session.add(lancamento)
        db.session.commit()
        flash('Lançamento registrado!', 'success')

    return redirect(url_for('main.lancamentos'))

@main.route('/lancamentos/excluir/<int:id>', methods=['POST'])
@login_required
def excluir_lancamento(id):
    lancamento = Lancamento.query.filter_by(id=id, usuario_id=current_user.id).first_or_404()
    db.session.delete(lancamento)
    db.session.commit()
    flash('Lançamento excluído.', 'info')
    return redirect(url_for('main.lancamentos'))

# ─── COFRES ──────────────────────────────────────────────────
@main.route('/cofres')
@login_required
def cofres():
    cofres = Cofre.query.filter_by(usuario_id=current_user.id).all()
    return render_template('cofres.html', cofres=cofres)

@main.route('/cofres/novo', methods=['POST'])
@login_required
def novo_cofre():
    cofre = Cofre(
        nome=request.form['nome'],
        meta=float(request.form['meta']),
        valor_atual=float(request.form.get('valor_atual', 0)),
        emoji=request.form.get('emoji', '🏦'),
        usuario_id=current_user.id
    )
    db.session.add(cofre)
    db.session.commit()
    flash('Cofre criado!', 'success')
    return redirect(url_for('main.cofres'))

@main.route('/cofres/depositar/<int:id>', methods=['POST'])
@login_required
def depositar_cofre(id):
    cofre = Cofre.query.filter_by(id=id, usuario_id=current_user.id).first_or_404()
    valor = float(request.form['valor'])
    acao = request.form.get('acao', 'depositar')
    if acao == 'retirar':
        cofre.valor_atual = max(0, cofre.valor_atual - valor)
        flash('Valor retirado do cofre.', 'info')
    else:
        cofre.valor_atual += valor
        flash('Valor depositado no cofre!', 'success')
    db.session.commit()
    return redirect(url_for('main.cofres'))

@main.route('/cofres/excluir/<int:id>', methods=['POST'])
@login_required
def excluir_cofre(id):
    cofre = Cofre.query.filter_by(id=id, usuario_id=current_user.id).first_or_404()
    db.session.delete(cofre)
    db.session.commit()
    flash('Cofre excluído.', 'info')
    return redirect(url_for('main.cofres'))

# ─── RELATÓRIO ───────────────────────────────────────────────
@main.route('/relatorio')
@login_required
def relatorio():
    mes = request.args.get('mes', date.today().month, type=int)
    ano = request.args.get('ano', date.today().year, type=int)

    lancamentos = Lancamento.query.filter_by(usuario_id=current_user.id).filter(
        db.extract('month', Lancamento.data) == mes,
        db.extract('year', Lancamento.data) == ano
    ).all()

    entradas = [l for l in lancamentos if l.tipo == 'entrada']
    saidas = [l for l in lancamentos if l.tipo == 'saida']

    por_categoria = {}
    for l in saidas:
        por_categoria[l.categoria] = por_categoria.get(l.categoria, 0) + l.valor

    total_entradas = sum(l.valor for l in entradas)
    total_saidas = sum(l.valor for l in saidas)

    return render_template('relatorio.html',
        mes=mes, ano=ano,
        entradas=entradas, saidas=saidas,
        total_entradas=total_entradas,
        total_saidas=total_saidas,
        por_categoria=por_categoria
    )

# ─── TRANSFERÊNCIAS ──────────────────────────────────────────
@main.route('/transferencias')
@login_required
def transferencias():
    contas = Conta.query.filter_by(usuario_id=current_user.id).all()
    cofres = Cofre.query.filter_by(usuario_id=current_user.id).all()
    historico = Transferencia.query.filter_by(usuario_id=current_user.id).order_by(Transferencia.data.desc()).all()
    return render_template('transferencias.html', contas=contas, cofres=cofres, historico=historico)

@main.route('/transferencias/nova', methods=['POST'])
@login_required
def nova_transferencia():
    tipo = request.form.get('tipo')
    valor = float(request.form.get('valor'))
    descricao = request.form.get('descricao', '')
    data = datetime.strptime(request.form.get('data'), '%Y-%m-%d').date()

    if tipo == 'conta_conta':
        origem_id = int(request.form.get('conta_origem_id'))
        destino_id = int(request.form.get('conta_destino_id'))
        origem = Conta.query.filter_by(id=origem_id, usuario_id=current_user.id).first_or_404()
        destino = Conta.query.filter_by(id=destino_id, usuario_id=current_user.id).first_or_404()
        db.session.add(Lancamento(descricao=f'Transferência para {destino.nome}', valor=valor, tipo='saida', categoria='Transferência', data=data, conta_id=origem_id, usuario_id=current_user.id))
        db.session.add(Lancamento(descricao=f'Transferência de {origem.nome}', valor=valor, tipo='entrada', categoria='Transferência', data=data, conta_id=destino_id, usuario_id=current_user.id))
        t = Transferencia(valor=valor, descricao=descricao or f'{origem.nome} → {destino.nome}', data=data, tipo='conta_conta', conta_origem_id=origem_id, conta_destino_id=destino_id, usuario_id=current_user.id)

    elif tipo == 'conta_cofre':
        conta_id = int(request.form.get('conta_id'))
        cofre_id = int(request.form.get('cofre_id'))
        Conta.query.filter_by(id=conta_id, usuario_id=current_user.id).first_or_404()
        cofre = Cofre.query.filter_by(id=cofre_id, usuario_id=current_user.id).first_or_404()
        db.session.add(Lancamento(descricao=f'Depósito no cofre {cofre.nome}', valor=valor, tipo='saida', categoria='Transferência', data=data, conta_id=conta_id, usuario_id=current_user.id))
        cofre.valor_atual += valor
        t = Transferencia(valor=valor, descricao=descricao or f'Conta → {cofre.nome}', data=data, tipo='conta_cofre', conta_origem_id=conta_id, cofre_id=cofre_id, usuario_id=current_user.id)

    elif tipo == 'cofre_conta':
        cofre_id = int(request.form.get('cofre_id'))
        conta_id = int(request.form.get('conta_id'))
        cofre = Cofre.query.filter_by(id=cofre_id, usuario_id=current_user.id).first_or_404()
        Conta.query.filter_by(id=conta_id, usuario_id=current_user.id).first_or_404()
        cofre.valor_atual = max(0, cofre.valor_atual - valor)
        db.session.add(Lancamento(descricao=f'Retirada do cofre {cofre.nome}', valor=valor, tipo='entrada', categoria='Transferência', data=data, conta_id=conta_id, usuario_id=current_user.id))
        t = Transferencia(valor=valor, descricao=descricao or f'{cofre.nome} → Conta', data=data, tipo='cofre_conta', conta_destino_id=conta_id, cofre_id=cofre_id, usuario_id=current_user.id)

    db.session.add(t)
    db.session.commit()
    flash('Transferência realizada!', 'success')
    return redirect(url_for('main.transferencias'))

# ─── CARTÕES DE CRÉDITO ──────────────────────────────────────
@main.route('/cartoes')
@login_required
def cartoes():
    cartoes = CartaoCredito.query.filter_by(usuario_id=current_user.id).all()
    return render_template('cartoes.html', cartoes=cartoes,
                           categorias=Lancamento.CATEGORIAS_SAIDA)

@main.route('/cartoes/novo', methods=['POST'])
@login_required
def novo_cartao():
    cartao = CartaoCredito(
        nome=request.form['nome'],
        limite=float(request.form.get('limite', 0)),
        dia_vencimento=int(request.form['dia_vencimento']),
        usuario_id=current_user.id
    )
    db.session.add(cartao)
    db.session.commit()
    flash('Cartão cadastrado!', 'success')
    return redirect(url_for('main.cartoes'))

@main.route('/cartoes/excluir/<int:id>', methods=['POST'])
@login_required
def excluir_cartao(id):
    cartao = CartaoCredito.query.filter_by(id=id, usuario_id=current_user.id).first_or_404()
    db.session.delete(cartao)
    db.session.commit()
    flash('Cartão excluído.', 'info')
    return redirect(url_for('main.cartoes'))

@main.route('/cartoes/<int:id>/compras')
@login_required
def compras_cartao(id):
    cartao = CartaoCredito.query.filter_by(id=id, usuario_id=current_user.id).first_or_404()
    mes = request.args.get('mes', date.today().month, type=int)
    ano = request.args.get('ano', date.today().year, type=int)
    compras = CompraCartao.query.filter_by(cartao_id=id, usuario_id=current_user.id).filter(
        db.extract('month', CompraCartao.data) == mes,
        db.extract('year', CompraCartao.data) == ano
    ).order_by(CompraCartao.data.desc()).all()
    total = sum(c.valor for c in compras)
    contas = Conta.query.filter_by(usuario_id=current_user.id).all()
    return render_template('compras_cartao.html', cartao=cartao, compras=compras,
                           total=total, mes=mes, ano=ano,
                           categorias=Lancamento.CATEGORIAS_SAIDA,
                           contas=contas)

@main.route('/cartoes/<int:id>/pagar-fatura', methods=['POST'])
@login_required
def pagar_fatura(id):
    cartao = CartaoCredito.query.filter_by(id=id, usuario_id=current_user.id).first_or_404()
    valor = float(request.form.get('valor', cartao.total_fatura_atual))
    conta_id = int(request.form.get('conta_id'))
    data = datetime.strptime(request.form.get('data'), '%Y-%m-%d').date()
    Conta.query.filter_by(id=conta_id, usuario_id=current_user.id).first_or_404()
    lancamento = Lancamento(
        descricao=f'Pagamento fatura {cartao.nome}',
        valor=valor,
        tipo='saida',
        categoria='Cartão de Crédito',
        data=data,
        conta_id=conta_id,
        usuario_id=current_user.id
    )
    db.session.add(lancamento)
    db.session.commit()
    flash(f'Fatura de R$ {"%.2f" % valor} paga com sucesso!', 'success')
    return redirect(url_for('main.compras_cartao', id=id))

@main.route('/cartoes/<int:id>/compras/nova', methods=['POST'])
@login_required
def nova_compra_cartao(id):
    CartaoCredito.query.filter_by(id=id, usuario_id=current_user.id).first_or_404()
    compra = CompraCartao(
        descricao=request.form['descricao'],
        valor=float(request.form['valor']),
        categoria=request.form['categoria'],
        data=datetime.strptime(request.form['data'], '%Y-%m-%d').date(),
        parcelas=int(request.form.get('parcelas', 1)),
        cartao_id=id,
        usuario_id=current_user.id
    )
    db.session.add(compra)
    db.session.commit()
    flash('Compra registrada!', 'success')
    return redirect(url_for('main.compras_cartao', id=id))

@main.route('/cartoes/compras/excluir/<int:id>', methods=['POST'])
@login_required
def excluir_compra(id):
    compra = CompraCartao.query.filter_by(id=id, usuario_id=current_user.id).first_or_404()
    cartao_id = compra.cartao_id
    db.session.delete(compra)
    db.session.commit()
    flash('Compra excluída.', 'info')
    return redirect(url_for('main.compras_cartao', id=cartao_id))

# ─── GASTOS FIXOS ────────────────────────────────────────────
@main.route('/gastos-fixos')
@login_required
def gastos_fixos():
    gastos = GastoFixo.query.filter_by(usuario_id=current_user.id).order_by(GastoFixo.dia_vencimento).all()
    total = sum(g.valor for g in gastos)
    total_pago = sum(g.valor for g in gastos if g.pago)
    total_pendente = sum(g.valor for g in gastos if not g.pago)
    return render_template('gastos_fixos.html', gastos=gastos,
                           total=total, total_pago=total_pago,
                           total_pendente=total_pendente,
                           categorias=Lancamento.CATEGORIAS_SAIDA)

@main.route('/gastos-fixos/novo', methods=['POST'])
@login_required
def novo_gasto_fixo():
    gasto = GastoFixo(
        nome=request.form['nome'],
        valor=float(request.form['valor']),
        categoria=request.form['categoria'],
        dia_vencimento=int(request.form['dia_vencimento']),
        usuario_id=current_user.id
    )
    db.session.add(gasto)
    db.session.commit()
    flash('Gasto fixo cadastrado!', 'success')
    return redirect(url_for('main.gastos_fixos'))

@main.route('/gastos-fixos/pagar/<int:id>', methods=['POST'])
@login_required
def pagar_gasto_fixo(id):
    gasto = GastoFixo.query.filter_by(id=id, usuario_id=current_user.id).first_or_404()
    gasto.pago = not gasto.pago
    db.session.commit()
    status = 'pago' if gasto.pago else 'pendente'
    flash(f'{gasto.nome} marcado como {status}.', 'success')
    return redirect(url_for('main.gastos_fixos'))

@main.route('/gastos-fixos/excluir/<int:id>', methods=['POST'])
@login_required
def excluir_gasto_fixo(id):
    gasto = GastoFixo.query.filter_by(id=id, usuario_id=current_user.id).first_or_404()
    db.session.delete(gasto)
    db.session.commit()
    flash('Gasto fixo excluído.', 'info')
    return redirect(url_for('main.gastos_fixos'))

# ─── IMPORTAR EXTRATO ────────────────────────────────────────
@main.route('/importar')
@login_required
def importar():
    contas = Conta.query.filter_by(usuario_id=current_user.id).all()
    cartoes = CartaoCredito.query.filter_by(usuario_id=current_user.id).all()
    return render_template('importar.html', contas=contas, cartoes=cartoes)

@main.route('/importar/processar', methods=['POST'])
@login_required
def processar_importacao():
    arquivo = request.files.get('arquivo')
    conta_id = request.form.get('conta_id')
    tipo_origem = request.form.get('tipo_origem')

    if not arquivo:
        flash('Nenhum arquivo enviado.', 'danger')
        return redirect(url_for('main.importar'))

    nome = arquivo.filename.lower()
    transacoes = []

    try:
        if nome.endswith('.csv'):
            conteudo = None
            for enc in ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']:
                try:
                    arquivo.seek(0)
                    conteudo = arquivo.read().decode(enc, errors='ignore')
                    break
                except Exception:
                    continue

            sep = ';' if conteudo.count(';') > conteudo.count(',') else ','
            linhas = conteudo.splitlines()

            # Encontra a linha do cabeçalho real procurando por palavras-chave de data/transação
            header_idx = 0
            palavras_header = ['date', 'data', 'release', 'transaction', 'descri', 'histor', 'memo', 'lancamento']
            for i, linha in enumerate(linhas):
                linha_lower = linha.lower()
                if any(p in linha_lower for p in palavras_header):
                    header_idx = i
                    break

            # Lê o CSV a partir do cabeçalho real
            conteudo_util = '\n'.join(linhas[header_idx:])
            reader = csv.DictReader(io.StringIO(conteudo_util), delimiter=sep)
            for row in reader:
                if any(v.strip() for v in row.values() if v):
                    transacoes.append(dict(row))

        elif nome.endswith('.xlsx'):
            try:
                import openpyxl
                arquivo.seek(0)
                wb = openpyxl.load_workbook(arquivo)
                ws = wb.active
                headers = []
                for c in ws[1]:
                    headers.append(str(c.value).strip() if c.value is not None else f'col_{c.column}')
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if any(v is not None for v in row):
                        transacoes.append(dict(zip(headers, [str(v) if v is not None else '' for v in row])))
            except ImportError:
                flash('Para importar XLSX instale openpyxl: pip install openpyxl', 'danger')
                return redirect(url_for('main.importar'))
        else:
            flash('Formato inválido. Use .csv ou .xlsx', 'danger')
            return redirect(url_for('main.importar'))

    except Exception as e:
        flash(f'Erro ao ler o arquivo: {str(e)}', 'danger')
        return redirect(url_for('main.importar'))

    if not transacoes:
        flash('Nenhuma transação encontrada no arquivo.', 'warning')
        return redirect(url_for('main.importar'))

    # Normaliza as transações detectando as colunas automaticamente
    import json
    from flask import session

    transacoes_normalizadas = []

    # Detecta os nomes das colunas uma vez
    if transacoes:
        colunas = list(transacoes[0].keys())

        # Detecta coluna de data
        col_data = None
        for c in colunas:
            if any(p in c.lower() for p in ['date', 'data', 'release', 'dt_']):
                col_data = c
                break

        # Detecta coluna de descrição
        col_desc = None
        for c in colunas:
            if any(p in c.lower() for p in ['descri', 'histor', 'memo', 'transaction_type', 'type', 'tipo', 'estabeleci', 'lancamento']):
                col_desc = c
                break

        # Detecta coluna de valor
        col_valor = None
        for c in colunas:
            if any(p in c.lower() for p in ['net_amount', 'valor', 'amount', 'value', 'quantia', 'vlr']):
                col_valor = c
                break
        # Fallback — pega a primeira coluna numérica
        if not col_valor:
            for c in colunas:
                if any(p in c.lower() for p in ['credit', 'debit', 'saldo', 'balance']):
                    col_valor = c
                    break

        for t in transacoes:
            try:
                # Descrição
                desc = ''
                if col_desc and t.get(col_desc):
                    desc = str(t[col_desc]).strip()
                if not desc:
                    desc = ' | '.join(str(v) for v in t.values() if v and str(v).strip())[:60]

                # Valor
                val_raw = str(t.get(col_valor, '') or '').strip()
                val_raw = val_raw.replace('R$', '').replace(' ', '')
                # Formato brasileiro: 1.234,56 → 1234.56
                if ',' in val_raw and '.' in val_raw:
                    val_raw = val_raw.replace('.', '').replace(',', '.')
                elif ',' in val_raw:
                    val_raw = val_raw.replace(',', '.')
                try:
                    valor = float(val_raw)
                except Exception:
                    valor = 0.0

                # Tipo — negativo = saída, positivo = entrada
                if valor < 0:
                    tipo = 'saida'
                    valor = abs(valor)
                else:
                    tipo = 'entrada'

                # Ignora linhas com valor zero
                if valor == 0:
                    continue

                # Detecta tipo pela descrição se valor for positivo
                desc_lower = desc.lower()
                if any(p in desc_lower for p in ['pagamento', 'compra', 'pix enviado', 'débito', 'debito', 'saque', 'retirado', 'reservado', 'uber', 'ifood']):
                    tipo = 'saida'
                elif any(p in desc_lower for p in ['pix recebido', 'crédito', 'credito', 'salário', 'salario', 'rendimento', 'venda', 'depósito', 'deposito']):
                    tipo = 'entrada'

                # Data
                data_raw = str(t.get(col_data, '') or '').strip()
                data_fmt = ''
                if data_raw:
                    # Tenta vários formatos
                    for fmt in ['%d-%m-%Y', '%d/%m/%Y', '%Y-%m-%d', '%m/%d/%Y', '%d-%m-%y', '%d/%m/%y']:
                        try:
                            from datetime import datetime as dt_parse
                            data_fmt = dt_parse.strptime(data_raw[:10], fmt).strftime('%Y-%m-%d')
                            break
                        except Exception:
                            continue

                # Categoria automática pela descrição
                categoria = 'Outros'
                if any(p in desc_lower for p in ['mercado', 'supermercado', 'alimenta', 'restaurante', 'lanche', 'ifood', 'rappi', 'padaria', 'açougue']):
                    categoria = 'Alimentação'
                elif any(p in desc_lower for p in ['uber', 'transporte', '99', 'posto', 'combustivel', 'onibus', 'metro', 'taxi']):
                    categoria = 'Transporte'
                elif any(p in desc_lower for p in ['aluguel', 'condominio', 'luz', 'agua', 'internet', 'energia', 'moradia']):
                    categoria = 'Moradia'
                elif any(p in desc_lower for p in ['farmacia', 'medico', 'hospital', 'saude', 'plano']):
                    categoria = 'Saúde'
                elif any(p in desc_lower for p in ['netflix', 'spotify', 'cinema', 'lazer', 'show', 'disney', 'youtube']):
                    categoria = 'Lazer'
                elif any(p in desc_lower for p in ['roupa', 'loja', 'shopping', 'zara', 'renner', 'riachuelo']):
                    categoria = 'Roupas'
                elif any(p in desc_lower for p in ['escola', 'faculdade', 'curso', 'educacao', 'livro', 'ensino']):
                    categoria = 'Educação'
                elif any(p in desc_lower for p in ['pix enviado', 'transferencia', 'ted', 'doc']):
                    categoria = 'Transferência para terceiros'
                elif tipo == 'entrada' and any(p in desc_lower for p in ['rendimento', 'salario', 'venda']):
                    categoria = 'Outros'

                transacoes_normalizadas.append({
                    'descricao': desc[:80],
                    'valor': round(valor, 2),
                    'tipo': tipo,
                    'data': data_fmt,
                    'categoria': categoria
                })

            except Exception:
                continue

    transacoes = transacoes_normalizadas if transacoes_normalizadas else transacoes

    session['transacoes_importadas'] = json.dumps(transacoes[:50], default=str)
    session['importar_conta_id'] = conta_id
    session['importar_tipo_origem'] = tipo_origem

    return redirect(url_for('main.revisar_importacao'))


@main.route('/importar/revisar')
@login_required
def revisar_importacao():
    import json
    from flask import session
    transacoes_raw = session.get('transacoes_importadas', '[]')
    transacoes = json.loads(transacoes_raw)
    conta_id = session.get('importar_conta_id')
    tipo_origem = session.get('importar_tipo_origem')
    contas = Conta.query.filter_by(usuario_id=current_user.id).all()
    cartoes = CartaoCredito.query.filter_by(usuario_id=current_user.id).all()
    return render_template('revisar_importacao.html',
                           transacoes=transacoes,
                           conta_id=conta_id,
                           tipo_origem=tipo_origem,
                           contas=contas,
                           cartoes=cartoes,
                           categorias_saida=Lancamento.CATEGORIAS_SAIDA,
                           categorias_entrada=Lancamento.CATEGORIAS_ENTRADA)


@main.route('/importar/confirmar', methods=['POST'])
@login_required
def confirmar_importacao():
    descricoes = request.form.getlist('descricao')
    valores = request.form.getlist('valor')
    tipos = request.form.getlist('tipo')
    categorias = request.form.getlist('categoria')
    datas = request.form.getlist('data')
    incluir = request.form.getlist('incluir')
    conta_id = int(request.form.get('conta_id'))
    tipo_origem = request.form.get('tipo_origem')

    count = 0
    for i in range(len(descricoes)):
        if str(i) not in incluir:
            continue
        try:
            valor = float(str(valores[i]).replace(',', '.').replace('R$', '').strip())
            if valor <= 0:
                continue
            data_obj = datetime.strptime(datas[i], '%Y-%m-%d').date()

            if tipo_origem == 'cartao':
                compra = CompraCartao(
                    descricao=descricoes[i],
                    valor=valor,
                    categoria=categorias[i],
                    data=data_obj,
                    parcelas=1,
                    cartao_id=conta_id,
                    usuario_id=current_user.id
                )
                db.session.add(compra)
            else:
                lancamento = Lancamento(
                    descricao=descricoes[i],
                    valor=valor,
                    tipo=tipos[i],
                    categoria=categorias[i],
                    data=data_obj,
                    conta_id=conta_id,
                    usuario_id=current_user.id
                )
                db.session.add(lancamento)
            count += 1
        except Exception:
            continue

    db.session.commit()
    flash(f'{count} transações importadas com sucesso!', 'success')
    return redirect(url_for('main.lancamentos'))